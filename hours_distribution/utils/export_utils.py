import os
from django.http import HttpResponse
from django.conf import settings
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import seaborn as sns
from io import BytesIO
import base64
from datetime import datetime

# Регистрируем кириллические шрифты для ReportLab
def register_cyrillic_fonts():
    """Регистрирует TrueType шрифты для поддержки кириллицы"""
    try:
        # Пытаемся использовать системные шрифты Windows
        font_paths = {
            'Arial': "C:\\Windows\\Fonts\\arial.ttf",
            'Arial_Bold': "C:\\Windows\\Fonts\\arialbd.ttf",
        }
        
        for font_name, font_path in font_paths.items():
            if os.path.exists(font_path):
                try:
                    pdfmetrics.registerFont(TTFont(font_name, font_path))
                    print(f"✓ Шрифт {font_name} успешно загружен")
                except Exception as e:
                    print(f"✗ Ошибка загрузки {font_name}: {e}")
    except Exception as e:
        print(f"Предупреждение: не удалось загрузить системные шрифты: {e}")

# Вызываем регистрацию шрифтов при импорте модуля
register_cyrillic_fonts()



class ExcelExporter:
    """Экспорт в Excel"""
    
    @staticmethod
    def export_department_report(department, workloads, year):
        """Экспорт отчета по кафедре"""
        wb = Workbook()
        ws = wb.active
        ws.title = department.short_name
        
        # Заголовок
        ws.merge_cells('A1:J1')
        ws['A1'] = f"Отчет по кафедре: {department.name}"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:J2')
        ws['A2'] = f"Учебный год: {year}"
        ws['A2'].font = Font(size=12)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Заголовки таблицы
        headers = ['№', 'Преподаватель', 'Дисциплина', 'Группа/Поток', 'Семестр', 'Курс', 'Лекции', 'Практики', 'Лаб.', 'Всего']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(bottom=Side(style='medium'))
        
        # Данные
        row = 5
        total_lecture = total_practice = total_lab = 0
        
        for i, wl in enumerate(workloads, 1):
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=str(wl.teacher))
            ws.cell(row=row, column=3, value=wl.curriculum_subject.subject.name)
            # Группа или поток
            csg = wl.curriculum_subject_group
            if csg and csg.student_group:
                group_display = csg.student_group.get_full_name()
            elif csg and csg.stream_group:
                group_display = f"Поток: {csg.stream_group.name}"
            else:
                group_display = ''
            ws.cell(row=row, column=4, value=group_display)
            # Семестр
            semester_val = getattr(wl.curriculum_subject.semester, 'number', '') if wl.curriculum_subject and wl.curriculum_subject.semester else ''
            ws.cell(row=row, column=5, value=semester_val)
            ws.cell(row=row, column=6, value=wl.curriculum_subject.course.number)
            ws.cell(row=row, column=7, value=wl.hours_lecture)
            ws.cell(row=row, column=8, value=wl.hours_practice)
            ws.cell(row=row, column=9, value=wl.hours_lab)
            ws.cell(row=row, column=10, value=wl.total_hours())
            
            total_lecture += wl.hours_lecture
            total_practice += wl.hours_practice
            total_lab += wl.hours_lab
            row += 1
        
        # Итоги
        row += 1
        ws.cell(row=row, column=6, value='ИТОГО:').font = Font(bold=True)
        ws.cell(row=row, column=7, value=total_lecture).font = Font(bold=True)
        ws.cell(row=row, column=8, value=total_practice).font = Font(bold=True)
        ws.cell(row=row, column=9, value=total_lab).font = Font(bold=True)
        ws.cell(row=row, column=10, value=total_lecture+total_practice+total_lab).font = Font(bold=True)
        
        # Настройка ширины столбцов
        column_widths = [5, 30, 40, 30, 10, 10, 10, 10, 10, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Сохранение в байтовый поток
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    @staticmethod
    def export_teacher_report(teacher, workloads, year):
        """Экспорт отчета по преподавателю"""
        # очень похож на экспорт по кафедре, но без колонки "Преподаватель"
        wb = Workbook()
        ws = wb.active
        ws.title = f"{teacher.last_name}_{teacher.first_name}"[:31]
        
        ws.merge_cells('A1:H1')
        ws['A1'] = f"Отчет по преподавателю: {teacher.last_name} {teacher.first_name}"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:H2')
        ws['A2'] = f"Учебный год: {year}"
        ws['A2'].font = Font(size=12)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        headers = ['№', 'Дисциплина', 'Группа/Поток', 'Семестр', 'Курс', 'Лекции', 'Практики', 'Лаб.', 'Всего']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(bottom=Side(style='medium'))
        
        row = 5
        total_lecture = total_practice = total_lab = 0
        for i, wl in enumerate(workloads, 1):
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=wl.curriculum_subject.subject.name)
            # Group/Stream
            csg = wl.curriculum_subject_group
            if csg and csg.student_group:
                group_display = csg.student_group.get_full_name()
            elif csg and csg.stream_group:
                group_display = f"Поток: {csg.stream_group.name}"
            else:
                group_display = ''
            ws.cell(row=row, column=3, value=group_display)
            # Semester
            semester_val = getattr(wl.curriculum_subject.semester, 'number', '') if wl.curriculum_subject and wl.curriculum_subject.semester else ''
            ws.cell(row=row, column=4, value=semester_val)
            ws.cell(row=row, column=5, value=wl.curriculum_subject.course.number)
            ws.cell(row=row, column=6, value=wl.hours_lecture)
            ws.cell(row=row, column=7, value=wl.hours_practice)
            ws.cell(row=row, column=8, value=wl.hours_lab)
            ws.cell(row=row, column=9, value=wl.total_hours())
            total_lecture += wl.hours_lecture
            total_practice += wl.hours_practice
            total_lab += wl.hours_lab
            row += 1
        
        row += 1
        ws.cell(row=row, column=5, value='ИТОГО:').font = Font(bold=True)
        ws.cell(row=row, column=6, value=total_lecture).font = Font(bold=True)
        ws.cell(row=row, column=7, value=total_practice).font = Font(bold=True)
        ws.cell(row=row, column=8, value=total_lab).font = Font(bold=True)
        ws.cell(row=row, column=9, value=total_lecture+total_practice+total_lab).font = Font(bold=True)

        column_widths = [5, 40, 30, 10, 8, 10, 10, 10, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def export_faculty_report(faculty, departments_data, year):
        """Экспорт отчета по факультету"""
        wb = Workbook()
        
        # Сводный лист
        ws_summary = wb.active
        ws_summary.title = "Сводный отчет"
        ws_summary.merge_cells('A1:D1')
        ws_summary['A1'] = f"Сводный отчет по факультету: {faculty.name}"
        ws_summary['A1'].font = Font(size=14, bold=True)
        ws_summary['A1'].alignment = Alignment(horizontal='center')
        
        headers = ['Кафедра', 'Преподавателей', 'Лекции', 'Практики', 'Лаб.', 'Всего часов']
        for col, header in enumerate(headers, 1):
            ws_summary.cell(row=3, column=col, value=header).font = Font(bold=True)
        
        for row, dept_data in enumerate(departments_data, start=4):
            ws_summary.cell(row=row, column=1, value=dept_data['department'].name)
            ws_summary.cell(row=row, column=2, value=dept_data['teachers_count'])
            ws_summary.cell(row=row, column=3, value=dept_data['lecture_hours'])
            ws_summary.cell(row=row, column=4, value=dept_data['practice_hours'])
            ws_summary.cell(row=row, column=5, value=dept_data['lab_hours'])
            ws_summary.cell(row=row, column=6, value=dept_data['total_hours'])
        
        # Листы по кафедрам
        for dept_data in departments_data:
            ws = wb.create_sheet(title=dept_data['department'].short_name[:31])
            ExcelExporter._add_department_sheet(ws, dept_data)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def _add_department_sheet(ws, dept_data):
        """Добавить лист с данными по кафедре"""
        ws.merge_cells('A1:H1')
        ws['A1'] = f"Кафедра: {dept_data['department'].name}"
        ws['A1'].font = Font(size=12, bold=True)
        
        headers = ['Преподаватель', 'Должность', 'Дисциплина', 'Группа/Поток', 'Семестр', 'Курс', 'Лекции', 'Практики', 'Лаб.', 'Всего']
        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header).font = Font(bold=True)
        
        row = 4
        for teacher_data in dept_data['teachers']:
            for workload in teacher_data['workloads']:
                ws.cell(row=row, column=1, value=str(teacher_data['teacher']))
                ws.cell(row=row, column=2, value=teacher_data['teacher'].position)
                ws.cell(row=row, column=3, value=workload.curriculum_subject.subject.name)
                # Group/Stream
                csg = workload.curriculum_subject_group
                if csg and csg.student_group:
                    group_display = csg.student_group.get_full_name()
                elif csg and csg.stream_group:
                    group_display = f"Поток: {csg.stream_group.name}"
                else:
                    group_display = ''
                ws.cell(row=row, column=4, value=group_display)
                # Semester
                semester_val = getattr(workload.curriculum_subject.semester, 'number', '') if workload.curriculum_subject and workload.curriculum_subject.semester else ''
                ws.cell(row=row, column=5, value=semester_val)
                ws.cell(row=row, column=6, value=workload.curriculum_subject.course.number)
                ws.cell(row=row, column=7, value=workload.hours_lecture)
                ws.cell(row=row, column=8, value=workload.hours_practice)
                ws.cell(row=row, column=9, value=workload.hours_lab)
                ws.cell(row=row, column=10, value=workload.total_hours())
                row += 1
    
    @staticmethod
    def export_departments_summary_report(department_stats, total_stats, year):
        """Экспорт итогового отчета по кафедрам с преподавателями"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Итоговый отчет"
        
        # Заголовок
        ws.merge_cells('A1:I1')
        ws['A1'] = f"Итоговый отчет по кафедрам - {year.name}"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Общая статистика
        ws['A3'] = "Общая статистика:"
        ws['A3'].font = Font(bold=True)
        
        stats_data = [
            ["Кафедр", total_stats['departments_count']],
            ["Преподавателей", total_stats['teachers_count']],
            ["Всего часов", total_stats['total_hours']],
            ["Лекционных", total_stats['lecture_hours']],
            ["Практических", total_stats['practice_hours']],
            ["Лабораторных", total_stats['lab_hours']],
        ]
        
        for row, (label, value) in enumerate(stats_data, start=4):
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
        
        # Таблица по кафедрам
        ws['A10'] = "Статистика по кафедрам:"
        ws['A10'].font = Font(bold=True)
        
        headers = ['№', 'Кафедра', 'Факультет', 'Преподавателей', 'Всего часов', 
                  'Лекционных', 'Практических', 'Лабораторных', 'Среднее на преподавателя']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=11, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(bottom=Side(style='medium'))
        
        for row, stat in enumerate(department_stats, start=12):
            ws.cell(row=row, column=1, value=row-11)
            ws.cell(row=row, column=2, value=stat['department'].name)
            ws.cell(row=row, column=3, value=stat['faculty'].short_name)
            ws.cell(row=row, column=4, value=stat['teachers_count'])
            ws.cell(row=row, column=5, value=stat['total_hours'])
            ws.cell(row=row, column=6, value=stat['lecture_hours'])
            ws.cell(row=row, column=7, value=stat['practice_hours'])
            ws.cell(row=row, column=8, value=stat['lab_hours'])
            ws.cell(row=row, column=9, value=round(stat['avg_hours_per_teacher'], 1))
        
        # Автоподбор ширины колонок
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Добавляем листы с преподавателями
        for dept_stat in department_stats:
            department = dept_stat['department']
            teachers = dept_stat.get('teachers', [])
            
            # Названия листов ограничены 31 символом
            sheet_name = f"{department.short_name}"[:31]
            ws_dept = wb.create_sheet(title=sheet_name)
            
            # Заголовок
            ws_dept['A1'] = f"Кафедра: {department.name}"
            ws_dept['A1'].font = Font(size=12, bold=True)
            
            # Таблица преподавателей
            headers_teachers = ['Фамилия', 'Имя', 'Лекции', 'Практики', 'Лабораторные', 'Всего часов']
            for col, header in enumerate(headers_teachers, 1):
                cell = ws_dept.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(bottom=Side(style='medium'))
            
            # Добавляем преподавателей
            if teachers:
                for row, teacher_stat in enumerate(teachers, start=4):
                    ws_dept.cell(row=row, column=1, value=teacher_stat['teacher'].last_name)
                    ws_dept.cell(row=row, column=2, value=teacher_stat['teacher'].first_name)
                    ws_dept.cell(row=row, column=3, value=teacher_stat['lecture_hours'])
                    ws_dept.cell(row=row, column=4, value=teacher_stat['practice_hours'])
                    ws_dept.cell(row=row, column=5, value=teacher_stat['lab_hours'])
                    ws_dept.cell(row=row, column=6, value=teacher_stat['total_hours'])
                
                # Итого по кафедре
                total_row = 4 + len(teachers)
                ws_dept.cell(row=total_row, column=1, value="ИТОГО")
                ws_dept.cell(row=total_row, column=1).font = Font(bold=True)
                ws_dept.cell(row=total_row, column=3, value=dept_stat['lecture_hours']).font = Font(bold=True)
                ws_dept.cell(row=total_row, column=4, value=dept_stat['practice_hours']).font = Font(bold=True)
                ws_dept.cell(row=total_row, column=5, value=dept_stat['lab_hours']).font = Font(bold=True)
                ws_dept.cell(row=total_row, column=6, value=dept_stat['total_hours']).font = Font(bold=True)
            else:
                ws_dept.cell(row=4, column=1, value="Нет преподавателей или нет нагрузки")
            
            # Автоподбор ширины
            for col in range(1, len(headers_teachers) + 1):
                ws_dept.column_dimensions[get_column_letter(col)].width = 15
        
        # Сохраняем в буфер
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def export_group_summary(groups_data, year):
        """Экспорт сводного отчета по группам"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Сводный отчет"
        
        # Заголовок
        ws.merge_cells('A1:J1')
        ws['A1'] = f"Сводный отчет по группам студентов - {year.name}"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Заголовки
        if groups_data:
            headers = list(groups_data[0].keys())
        else:
            headers = ['Код группы', 'Название', 'Тип', 'Курс', 'Направление', 
                      'Факультет', 'Год набора', 'Кол-во студентов', 
                      'Всего часов', 'Часов на студента']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(bottom=Side(style='medium'))
        
        for row, data in enumerate(groups_data, start=4):
            for col, (key, value) in enumerate(data.items(), 1):
                ws.cell(row=row, column=col, value=value)
        
        # Настройка ширины столбцов
        column_widths = [15, 25, 12, 8, 12, 15, 10, 12, 12, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def export_group_detailed(workloads, year):
        """Экспорт детального отчета по группам"""
        wb = Workbook()
        
        # Основной лист
        ws_main = wb.active
        ws_main.title = "Нагрузка по группам"
        
        ws_main.merge_cells('A1:I1')
        ws_main['A1'] = f"Детальный отчет по нагрузке - {year.name}"
        ws_main['A1'].font = Font(size=14, bold=True)
        ws_main['A1'].alignment = Alignment(horizontal='center')
        
        headers = ['Группа/Поток', 'Тип группы', 'Дисциплина', 'Код дисциплины', 'Семестр',
              'Преподаватель', 'Лекции', 'Практики', 'Лаб.', 'Всего часов']
        
        for col, header in enumerate(headers, 1):
            ws_main.cell(row=3, column=col, value=header).font = Font(bold=True)
            ws_main.cell(row=3, column=col).alignment = Alignment(horizontal='center')
        
        row = 4
        for wl in workloads:
            csg = wl.curriculum_subject_group
            subject = csg.curriculum_subject.subject
            
            if csg.student_group:
                group_name = csg.student_group.get_full_name()
                group_type = 'Обычная'
            else:
                group_name = f"Поток: {csg.stream_group.name}"
                group_type = 'Потоковая'
            
            ws_main.cell(row=row, column=1, value=group_name)
            ws_main.cell(row=row, column=2, value=group_type)
            ws_main.cell(row=row, column=3, value=subject.name)
            ws_main.cell(row=row, column=4, value=subject.code)
            # Semester
            sem = getattr(csg.curriculum_subject.semester, 'number', '') if csg.curriculum_subject and csg.curriculum_subject.semester else ''
            ws_main.cell(row=row, column=5, value=sem)
            ws_main.cell(row=row, column=6, value=str(wl.teacher))
            ws_main.cell(row=row, column=7, value=wl.hours_lecture)
            ws_main.cell(row=row, column=8, value=wl.hours_practice)
            ws_main.cell(row=row, column=9, value=wl.hours_lab)
            ws_main.cell(row=row, column=10, value=wl.total_hours())
            row += 1
        
        # Лист статистики
        ws_stats = wb.create_sheet(title="Статистика")
        
        # Группировка по типам
        group_stats = {}
        for wl in workloads:
            csg = wl.curriculum_subject_group
            if csg.student_group:
                key = 'Обычные группы'
                students_count = csg.student_group.students_count
            else:
                key = 'Потоковые группы'
                students_count = csg.stream_group.total_students
            
            if key not in group_stats:
                group_stats[key] = {
                    'count': 0,
                    'total_hours': 0,
                    'total_students': 0
                }
            
            stats = group_stats[key]
            stats['count'] += 1
            stats['total_hours'] += wl.total_hours()
            stats['total_students'] += students_count
        
        ws_stats['A1'] = 'Статистика по типам групп'
        ws_stats['A1'].font = Font(bold=True)
        
        headers_stats = ['Тип группы', 'Кол-во записей', 'Всего часов', 'Всего студентов', 'Часов на студента']
        for col, header in enumerate(headers_stats, 1):
            ws_stats.cell(row=3, column=col, value=header).font = Font(bold=True)
        
        for row, (type_name, stats) in enumerate(group_stats.items(), start=4):
            ws_stats.cell(row=row, column=1, value=type_name)
            ws_stats.cell(row=row, column=2, value=stats['count'])
            ws_stats.cell(row=row, column=3, value=stats['total_hours'])
            ws_stats.cell(row=row, column=4, value=stats['total_students'])
            if stats['total_students'] > 0:
                hours_per_student = stats['total_hours'] / stats['total_students']
                ws_stats.cell(row=row, column=5, value=round(hours_per_student, 2))
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
class PDFExporter:
    """Экспорт в PDF"""
    
    @staticmethod
    def export_department_report(department, workloads, year):
        """Экспорт отчета по кафедре в PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        
        styles = getSampleStyleSheet()
        
        # Стиль заголовка с поддержкой кириллицы
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1,  # center
            fontName='Arial' if 'Arial' in pdfmetrics.getRegisteredFontNames() else 'Helvetica',
        )
        
        # Стиль для обычного текста
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=11,
            fontName='Arial' if 'Arial' in pdfmetrics.getRegisteredFontNames() else 'Helvetica',
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=12,
            fontName='Arial' if 'Arial' in pdfmetrics.getRegisteredFontNames() else 'Helvetica',
        )
        
        content = []
        
        # Заголовок
        content.append(Paragraph(f"Отчет по кафедре: {department.name}", title_style))
        content.append(Paragraph(f"Учебный год: {year}", heading_style))
        content.append(Spacer(1, 20))
        
        # Таблица (добавлены колонки: Группа/Поток, Семестр)
        data = [['№', 'Преподаватель', 'Дисциплина', 'Группа/Поток', 'Семестр', 'Курс', 'Лекции', 'Практики', 'Лаб.', 'Всего']]

        total_lecture = total_practice = total_lab = 0

        for i, wl in enumerate(workloads, 1):
            csg = wl.curriculum_subject_group
            if csg and csg.student_group:
                group_display = csg.student_group.get_full_name()
            elif csg and csg.stream_group:
                group_display = f"Поток: {csg.stream_group.name}"
            else:
                group_display = ''

            semester_val = ''
            if wl.curriculum_subject and getattr(wl.curriculum_subject, 'semester', None):
                semester_val = getattr(wl.curriculum_subject.semester, 'number', '')

            data.append([
                str(i),
                str(wl.teacher),
                wl.curriculum_subject.subject.name,
                group_display,
                str(semester_val),
                str(wl.curriculum_subject.course.number),
                str(wl.hours_lecture),
                str(wl.hours_practice),
                str(wl.hours_lab),
                str(wl.total_hours())
            ])
            total_lecture += wl.hours_lecture
            total_practice += wl.hours_practice
            total_lab += wl.hours_lab

        # Итоговая строка
        data.append(['', '', '', '', 'ИТОГО:', str(total_lecture), str(total_practice), str(total_lab), str(total_lecture + total_practice + total_lab), ''])

        table = Table(data, colWidths=[2*cm, 5*cm, 8*cm, 6*cm, 2*cm, 2*cm, 2*cm, 2*cm, 2*cm, 2*cm])
        
        # Определяем шрифты
        font_name = 'Arial' if 'Arial' in pdfmetrics.getRegisteredFontNames() else 'Helvetica'
        font_bold = 'Arial_Bold' if 'Arial_Bold' in pdfmetrics.getRegisteredFontNames() else 'Helvetica-Bold'
        
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), font_bold),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), font_bold),
            ('FONTNAME', (0, 1), (-1, -2), font_name),
        ]))
        
        content.append(table)
        doc.build(content)
        buffer.seek(0)
        return buffer

    @staticmethod
    def export_teacher_report(teacher, workloads, year):
        """Экспорт отчета по преподавателю в PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1,
            fontName='Arial' if 'Arial' in pdfmetrics.getRegisteredFontNames() else 'Helvetica',
        )
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=11,
            fontName='Arial' if 'Arial' in pdfmetrics.getRegisteredFontNames() else 'Helvetica',
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=12,
            fontName='Arial' if 'Arial' in pdfmetrics.getRegisteredFontNames() else 'Helvetica',
        )
        content = []
        content.append(Paragraph(f"Отчет по преподавателю: {teacher.last_name} {teacher.first_name}", title_style))
        content.append(Paragraph(f"Учебный год: {year}", heading_style))
        content.append(Spacer(1, 20))
        data = [['№', 'Дисциплина', 'Группа/Поток', 'Семестр', 'Курс', 'Лекции', 'Практики', 'Лаб.', 'Всего']]
        total_lecture = total_practice = total_lab = 0
        for i, wl in enumerate(workloads, 1):
            csg = wl.curriculum_subject_group
            if csg and csg.student_group:
                group_display = csg.student_group.get_full_name()
            elif csg and csg.stream_group:
                group_display = f"Поток: {csg.stream_group.name}"
            else:
                group_display = ''

            semester_val = ''
            if wl.curriculum_subject and getattr(wl.curriculum_subject, 'semester', None):
                semester_val = getattr(wl.curriculum_subject.semester, 'number', '')

            data.append([
                str(i),
                wl.curriculum_subject.subject.name,
                group_display,
                str(semester_val),
                str(wl.curriculum_subject.course.number),
                str(wl.hours_lecture),
                str(wl.hours_practice),
                str(wl.hours_lab),
                str(wl.total_hours())
            ])
            total_lecture += wl.hours_lecture
            total_practice += wl.hours_practice
            total_lab += wl.hours_lab
        data.append(['', '', '', 'ИТОГО:', str(total_lecture), str(total_practice), str(total_lab), str(total_lecture + total_practice + total_lab), ''])
        table = Table(data, colWidths=[2*cm, 8*cm, 6*cm, 2*cm, 2*cm, 2*cm, 2*cm, 2*cm, 2*cm])
        font_name = 'Arial' if 'Arial' in pdfmetrics.getRegisteredFontNames() else 'Helvetica'
        font_bold = 'Arial_Bold' if 'Arial_Bold' in pdfmetrics.getRegisteredFontNames() else 'Helvetica-Bold'
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), font_bold),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), font_bold),
            ('FONTNAME', (0, 1), (-1, -2), font_name),
        ]))
        content.append(table)
        doc.build(content)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def export_departments_summary_report(department_stats, total_stats, year):
        """Экспорт итогового отчета по кафедрам в PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        
        styles = getSampleStyleSheet()
        
        # Стиль заголовка с поддержкой кириллицы
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1,  # center
            fontName='Arial' if 'Arial' in pdfmetrics.getRegisteredFontNames() else 'Helvetica',
        )
        
        # Стиль для обычного текста
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            fontName='Arial' if 'Arial' in pdfmetrics.getRegisteredFontNames() else 'Helvetica',
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=12,
            fontName='Arial' if 'Arial' in pdfmetrics.getRegisteredFontNames() else 'Helvetica',
        )
        
        content = []
        
        # Заголовок
        content.append(Paragraph(f"Итоговый отчет по кафедрам - {year.name}", title_style))
        content.append(Spacer(1, 12))
        
        # Общая статистика
        content.append(Paragraph("Общая статистика:", heading_style))
        content.append(Spacer(1, 6))
        
        stats_text = f"""
        Кафедр: {total_stats['departments_count']}<br/>
        Преподавателей: {total_stats['teachers_count']}<br/>
        Всего часов: {total_stats['total_hours']}<br/>
        Лекционных: {total_stats['lecture_hours']}<br/>
        Практических: {total_stats['practice_hours']}<br/>
        Лабораторных: {total_stats['lab_hours']}
        """
        content.append(Paragraph(stats_text, normal_style))
        content.append(Spacer(1, 12))
        
        # Таблица по кафедрам
        content.append(Paragraph("Статистика по кафедрам:", heading_style))
        content.append(Spacer(1, 6))
        
        table_data = [['№', 'Кафедра', 'Факультет', 'Преподавателей', 'Всего часов', 
                      'Лекционных', 'Практических', 'Лабораторных']]
        
        for i, stat in enumerate(department_stats, 1):
            table_data.append([
                str(i),
                stat['department'].name[:20],
                stat['faculty'].short_name,
                str(stat['teachers_count']),
                str(stat['total_hours']),
                str(stat['lecture_hours']),
                str(stat['practice_hours']),
                str(stat['lab_hours'])
            ])
        
        font_name = 'Arial' if 'Arial' in pdfmetrics.getRegisteredFontNames() else 'Helvetica'
        font_bold = 'Arial_Bold' if 'Arial_Bold' in pdfmetrics.getRegisteredFontNames() else 'Helvetica-Bold'
        table = Table(table_data, colWidths=[1*cm, 3*cm, 1.5*cm, 1.5*cm, 1.5*cm, 1.5*cm, 1.5*cm, 1.5*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), font_bold),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), font_name),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        content.append(table)
        content.append(Spacer(1, 20))
        
        # Таблица преподавателей для каждой кафедры
        for dept_stat in department_stats:
            department = dept_stat['department']
            teachers = dept_stat['teachers']
            
            content.append(Paragraph(f"Кафедра: {department.name}", heading_style))
            content.append(Spacer(1, 6))
            
            if teachers:
                teachers_table_data = [['Фамилия', 'Имя', 'Лекции', 'Практики', 'Лабораторные', 'Всего']]
                
                for teacher_stat in teachers:
                    teachers_table_data.append([
                        teacher_stat['teacher'].last_name,
                        teacher_stat['teacher'].first_name,
                        str(teacher_stat['lecture_hours']),
                        str(teacher_stat['practice_hours']),
                        str(teacher_stat['lab_hours']),
                        str(teacher_stat['total_hours'])
                    ])
                
                # Итого по кафедре
                teachers_table_data.append([
                    'ИТОГО',
                    '',
                    str(dept_stat['lecture_hours']),
                    str(dept_stat['practice_hours']),
                    str(dept_stat['lab_hours']),
                    str(dept_stat['total_hours'])
                ])
                
                teachers_table = Table(teachers_table_data, colWidths=[2.5*cm, 2.5*cm, 1.5*cm, 1.5*cm, 1.5*cm, 1.5*cm])
                teachers_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), font_bold),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('GRID', (0, 0), (-1, -2), 1, colors.black),
                    ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                    ('FONTNAME', (0, -1), (-1, -1), font_bold),
                    ('FONTNAME', (0, 1), (-1, -2), font_name),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                ]))
                content.append(teachers_table)
            else:
                content.append(Paragraph("Нет преподавателей или нет нагрузки", normal_style))
            
            content.append(Spacer(1, 12))
        
        doc.build(content)
        buffer.seek(0)
        return buffer

class ChartGenerator:
    """Генератор графиков и диаграмм"""
    
    @staticmethod
    def generate_pie_chart(labels, sizes, title):
        """Создать круговую диаграмму"""
        plt.figure(figsize=(8, 6))
        plt.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90)
        plt.axis('equal')
        plt.title(title)
        
        buffer = BytesIO()
        plt.savefig(buffer, format='png', dpi=100, bbox_inches='tight')
        buffer.seek(0)
        plt.close()
        
        image_base64 = base64.b64encode(buffer.getvalue()).decode()
        return f"data:image/png;base64,{image_base64}"
    
    @staticmethod
    def generate_bar_chart(labels, values, title, xlabel, ylabel):
        """Создать столбчатую диаграмму"""
        plt.figure(figsize=(10, 6))
        bars = plt.bar(labels, values)
        plt.title(title)
        plt.xlabel(xlabel)
        plt.ylabel(ylabel)
        plt.xticks(rotation=45, ha='right')
        
        # Добавить значения на столбцы
        for bar in bars:
            height = bar.get_height()
            plt.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                    f'{height:.0f}', ha='center', va='bottom')
        
        plt.tight_layout()
        
        buffer = BytesIO()
        plt.savefig(buffer, format='png', dpi=100)
        buffer.seek(0)
        plt.close()
        
        image_base64 = base64.b64encode(buffer.getvalue()).decode()
        return f"data:image/png;base64,{image_base64}"
    
    # Создать набор графиков для факультета
    @staticmethod
    def generate_faculty_charts(faculty_data):
        charts = {}
        
        # Круговая диаграмма распределения часов по кафедрам
        dept_names = [dept['department'].short_name for dept in faculty_data]
        dept_hours = [dept['total_hours'] for dept in faculty_data]
        
        charts['departments_pie'] = ChartGenerator.generate_pie_chart(
            dept_names, dept_hours, 'Распределение часов по кафедрам'
        )
        
        # Столбчатая диаграмма преподавателей по кафедрам
        dept_teachers = [dept['teachers_count'] for dept in faculty_data]
        
        charts['teachers_bar'] = ChartGenerator.generate_bar_chart(
            dept_names, dept_teachers,
            'Количество преподавателей по кафедрам',
            'Кафедры', 'Преподаватели'
        )
        
        return charts